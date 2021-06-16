import pandas as pd
import numpy as np
from qmin import load_model
from qmin import load_regression_model

class MolWeight:

    def __init__(self):
        self.SIO2 = 60.0843
        self.TIO2 = 79.8988
        self.AL2O3 = 101.9613
        self.CR2O3 = 159.6922
        self.FEO = 71.6464
        self.MNO = 70.9374
        self.CAO = 56.0794
        self.MGO = 40.3044
        self.NA2O = 61.9794
        self.K2O = 94.196
        self.F = 18.998
        self.LI2O = 29.887
        self.CL = 35.453
        self.ZNO = 81.38


def organize(df):
    model = load_model()
    table_reference = './assets/OXIDE_TO_ELEMENT.csv'
    df_references = pd.read_csv(table_reference, sep=';')

    dic = {}
    el = []
    for i in range(len(df_references['Element'])):
        dic[i] = str(df_references['Element'][i]).strip().upper()
        el.append(str(df_references['Element'][i]).strip().upper())
    df_references.replace(dic)
    df_references['Element_Upper'] = el

    # Loop to convert Element to oxide if needed
    # TODO: Check for + in ion exp Ca2+

    for i in df.columns:
        c = i.strip().upper()

        a = df_references[df_references['Element_Upper'] == c]

        if len(a) != 0:
            df[i] = df[i].astype('float')
            if c == 'S' or c == 'AG' or c == 'AS' or i == 'F' or c == 'CL':
                continue
            else:
                if len(a[a['Valency'] == 1]['Factor']) == 1:
                    if c == 'CU':
                        factor = float(a[a['Valency'] == 2]['Factor'])

                        df[c + 'O'] = df[i] * factor
                    else:
                        factor = float(a[a['Valency'] == 1]['Factor'])

                        df[c + 'O'] = df[i] * factor

                elif len(a[a['Valency'] == 2]['Factor']) == 1:
                    factor = float(a[a['Valency'] == 2]['Factor'])
                    df[c + 'O'] = df[i] * factor

                elif len(a[a['Valency'] == 3]['Factor']) == 1:
                    factor = float(a[a['Valency'] == 3]['Factor'])

                    if c == 'CO':
                        df[c + 'O'] = df[i] * factor
                    else:
                        df[c + '2O3'] = df[i] * factor

                elif len(a[a['Valency'] == 5]['Factor']) == 1:
                    factor = float(a[a['Valency'] == 5]['Factor'])
                    df[c + '2O5'] = df[i] * factor

    dic = {}
    for i in range(len(df.columns)):
        dic[df.columns[i]] = str(df.columns[i]).strip().upper()
    df = df.rename(columns=dic)

    for i in df.columns:
        if i == 'FEOT':
            df = df.rename(columns={'FEOT': 'FEO'})
            continue
    # Add missing columns
    for i in model['Train Features']:
        if i not in df.columns:
            df[i] = 0

    return df


def get_weights(df_grt):
    # Return DataFrame of moles weights for use in formulas

    weight = MolWeight()
    oxides = df_grt.columns
    # print(oxides)

    dic = dict()
    dic['SI_mole'] = df_grt['SIO2'] * 1 / weight.SIO2
    dic['SI_Oxygen'] = df_grt['SIO2'] * 2 / weight.SIO2

    dic['TI_mole'] = df_grt['TIO2'] * 1 / weight.TIO2
    dic['TI_Oxygen'] = df_grt['TIO2'] * 2 / weight.TIO2

    dic['AL_mole'] = df_grt['AL2O3'] * 2 / weight.AL2O3
    dic['AL_Oxygen'] = df_grt['AL2O3'] * 3 / weight.AL2O3

    dic['CR_mole'] = df_grt['CR2O3'] * 2 / weight.CR2O3
    dic['CR_Oxygen'] = df_grt['CR2O3'] * 3 / weight.CR2O3

    if 'FEO' in oxides:
        dic['FE_mole'] = df_grt['FEO'] * 1 / weight.FEO
        dic['FE_Oxygen'] = df_grt['FEO'] * 1 / weight.FEO
    else:
        dic['FE_mole'] = df_grt['FEOT'] * 1 / weight.FEO
        dic['FE_Oxygen'] = df_grt['FEOT'] * 1 / weight.FEO

    dic['MN_mole'] = df_grt['MNO'] * 1 / weight.MNO
    dic['MN_Oxygen'] = df_grt['MNO'] * 1 / weight.MNO

    dic['MG_mole'] = df_grt['MGO'] * 1 / weight.MGO
    dic['MG_Oxygen'] = df_grt['MGO'] * 1 / weight.MGO

    dic['CA_mole'] = df_grt['CAO'] * 1 / weight.CAO
    dic['CA_Oxygen'] = df_grt['CAO'] * 1 / weight.CAO

    if 'NA2O' in oxides:
        dic['NA_mole'] = df_grt['NA2O'] * 2 / weight.NA2O
        dic['NA_Oxygen'] = df_grt['NA2O'] * 1 / weight.NA2O
    if 'K2O' in oxides:
        dic['K_mole'] = df_grt['K2O'] * 2 / weight.K2O
        dic['K_Oxygen'] = df_grt['K2O'] * 1 / weight.K2O
    if 'CL' in oxides:
        dic['CL_mole'] = df_grt['CL'] * 1 / weight.CL
    if 'F' in oxides:
        dic['F_mole'] = df_grt['F'] * 1 / weight.CL
    if 'ZNO' in oxides:
        dic['ZN_mole'] = df_grt['ZNO'] * 2 / weight.ZNO
        dic['ZN_Oxygen'] = df_grt['ZNO'] * 1 / weight.ZNO

    df_f = pd.DataFrame(data=dic)
    df_f['moles_cation'] = df_f.filter(regex='_mole$', axis=1).sum(axis=1)
    df_f['moles_oxygen'] = df_f.filter(regex='_Oxygen$', axis=1).sum(axis=1)

    return df_f


def _norm_cations_oxygen(df_f, oxides, cation_multiplier):
    for eli in oxides:
        el = eli.split('O')[0]
        if len(el) > 2:
            el = el[:2]
        if el[0] == 'K':
            el = 'K'

        el = el.upper()

        df_f[el + '_NormCations'] = cation_multiplier * df_f[el + '_mole'] / df_f['moles_cation']
        if el in ['SI', 'TI']:
            df_f[el + '_NormOxygen'] = df_f[el + '_NormCations'] * 2
            df_f[el + '_Atom'] = df_f[el + '_NormCations'].round(1)
        elif el in ['AL', 'CR']:
            df_f[el + '_NormOxygen'] = df_f[el + '_NormCations'] * 3 / 2
            df_f[el + '_Atom'] = df_f[el + '_NormCations'].round(1)
        elif el in ['FE', 'MN', 'MG', 'CA']:
            df_f[el + '_NormOxygen'] = df_f[el + '_NormCations']
            df_f[el + '_Atom'] = df_f[el + '_NormCations'].round(1)
        elif el in ['K', 'NA', 'ZN']:
            df_f[el + '_NormOxygen'] = df_f[el + '_NormCations'] / 2
            df_f[el + '_Atom'] = df_f[el + '_NormCations'].round(1)

    df_f['NormCations_Total'] = df_f.filter(regex='_NormCations$', axis=1).sum(axis=1)
    df_f['NormOxygen_Total'] = df_f.filter(regex='_NormOxygen$', axis=1).sum(axis=1)

    df_f['calculate_charge'] = 4 * (df_f['SI_NormCations'] + df_f['TI_NormCations']) + \
                               3 * (df_f['AL_NormCations'] + df_f['CR_NormCations']) + \
                               2 * (df_f['FE_NormCations'] + df_f['CA_NormCations']) + \
                               2 * (df_f['MN_NormCations'] + df_f['MG_NormCations'])
    return df_f


def _test_Fe(df_f, value):
    df_f.loc[df_f['calculate_charge'] - value < 0, 'FE3_Atom'] = value - df_f['calculate_charge']
    df_f.loc[df_f['calculate_charge'] - value >= 0, 'FE3_Atom'] = 0

    df_f.loc[df_f['FE_NormCations'] - df_f['FE3_Atom'] > 0, 'FE2_Atom'] = df_f['FE_NormCations'] - df_f['FE3_Atom']
    df_f.loc[df_f['FE_NormCations'] - df_f['FE3_Atom'] <= 0, 'FE2_Atom'] = 0

    return df_f


def garnet_formula(df_grt):
    df_f = get_weights(df_grt)
    # Oxides of Interest for garnet
    oxides = ['SiO2', 'TiO2', 'Al2O3', 'Cr2O3', 'FeO', 'MnO', 'MgO', 'CaO']

    df_f = _norm_cations_oxygen(df_f, oxides, 8)
    df_f = _test_Fe(df_f, 24)

    df_f['almandine'] = np.round(100 * (df_f['FE2_Atom'] / (df_f['FE2_Atom'] + df_f['MN_Atom'] + df_f['MG_Atom']
                                                            + df_f['CA_Atom'])), 2)
    df_f['pyrope'] = np.round(100 * (df_f['MG_Atom'] / (df_f['FE2_Atom'] + df_f['MN_Atom'] + df_f['MG_Atom']
                                                        + df_f['CA_Atom'])), 2)
    df_f['grossular'] = np.round(100 * (df_f['CA_Atom'] / (df_f['FE2_Atom'] + df_f['MN_Atom'] + df_f['MG_Atom']
                                                           + df_f['CA_Atom'])) *
                                 ((df_f['AL_Atom'] / (df_f['TI_Atom'] + df_f['AL_Atom'] + df_f['CR_Atom']
                                                      + df_f['FE3_Atom']))), 2)
    df_f['spessartine'] = np.round(100 * (df_f['MN_Atom'] / (df_f['FE2_Atom'] + df_f['MN_Atom'] + df_f['MG_Atom']
                                                             + df_f['CA_Atom'])), 2)
    df_f['uvarovite'] = np.round(100 * (df_f['CR_Atom'] / (df_f['FE2_Atom'] + df_f['MN_Atom'] + df_f['MG_Atom']
                                                           + df_f['CA_Atom'])) *
                                 ((df_f['CA_Atom'] / (df_f['TI_Atom'] + df_f['AL_Atom'] + df_f['CR_Atom']
                                                      + df_f['FE3_Atom']))), 2)
    df_f['andradite'] = np.round(100 * (df_f['FE3_Atom'] / (df_f['FE2_Atom'] + df_f['MN_Atom'] + df_f['MG_Atom']
                                                            + df_f['CA_Atom'])) *
                                 ((df_f['CA_Atom'] / (df_f['TI_Atom'] + df_f['AL_Atom'] + df_f['CR_Atom']
                                                      + df_f['FE3_Atom']))), 2)
    df_f['CaTi_Grt'] = np.round(100 * (df_f['TI_Atom'] / (df_f['FE2_Atom'] + df_f['MN_Atom'] + df_f['MG_Atom']
                                                          + df_f['CA_Atom'])) *
                                ((df_f['CA_Atom'] / (df_f['TI_Atom'] + df_f['AL_Atom'] + df_f['CR_Atom']
                                                     + df_f['FE3_Atom']))), 2)
    df_f['Mol_Total'] = df_f['CaTi_Grt'] + df_f['andradite'] + df_f['uvarovite'] + df_f['spessartine'] + df_f[
        'grossular'] \
                        + df_f['pyrope'] + df_f['almandine']

    df_classification = df_f[['CaTi_Grt', 'andradite', 'uvarovite', 'spessartine', 'grossular', 'pyrope', 'almandine']]
    df_f['Classification'] = df_classification.idxmax(axis=1)

    # Round

    df_f['FE2_Atom'] = df_f['FE2_Atom'].round(1)
    df_f['FE3_Atom'] = df_f['FE3_Atom'].round(1)

    # Formula Garnet X3y2(TO4)3
    # cubico
    df_f['Formula'] = '('
    df_f.loc[df_f['CA_Atom'] > 0, 'Formula'] += 'Ca' + df_f['CA_Atom'].astype('str')
    df_f.loc[df_f['MN_Atom'] > 0, 'Formula'] += 'Mn' + df_f['MN_Atom'].astype('str')
    df_f.loc[df_f['FE2_Atom'] > 0, 'Formula'] += 'Fe' + df_f['FE2_Atom'].astype('str')
    df_f.loc[df_f['MG_Atom'] > 0, 'Formula'] += 'Mg' + df_f['MG_Atom'].astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += (df_f['CA_Atom'] + df_f['FE2_Atom'] + df_f['MN_Atom'] + df_f['MG_Atom']).round(1).astype('str')

    # octaedrico
    df_f['Formula'] += '('
    df_f.loc[df_f['TI_Atom'] > 0, 'Formula'] += 'Ti' + df_f['TI_Atom'].astype('str')
    df_f.loc[df_f['CR_Atom'] > 0, 'Formula'] += 'Cr' + df_f['CR_Atom'].astype('str')
    df_f.loc[df_f['FE3_Atom'] > 0, 'Formula'] += 'Fe' + df_f['FE3_Atom'].astype('str')
    df_f.loc[df_f['AL_Atom'] > 0, 'Formula'] += 'Al' + df_f['AL_Atom'].astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += (df_f['FE3_Atom'] + df_f['AL_Atom'] + df_f['CR_Atom'] + df_f['TI_Atom']).round(1).astype('str')

    # tetraedrico
    df_f['Formula'] += '('
    df_f.loc[df_f['SI_Atom'] > 0, 'Formula'] += 'Si' + df_f['SI_Atom'].astype('str')
    df_f['Formula'] += 'O' + (df_f['SI_NormOxygen'] / df_f['SI_Atom']).round(1).astype('str') + ')'
    df_f['Formula'] += '\u03A3' + (df_f['SI_Atom']).round(1).astype('str')

    # print(df_f['Formula'])
    # df_f.to_excel('formula_garnet.xlsx')

    return df_f[['Formula', 'CaTi_Grt', 'andradite', 'uvarovite', 'spessartine',
                 'grossular', 'pyrope', 'almandine', 'SI_Atom', 'TI_Atom',
                 'CR_Atom', 'AL_Atom', 'FE3_Atom', 'FE2_Atom', 'MN_Atom',
                 'CA_Atom', 'MG_Atom']]


def feldspar_formula(df_fdp):
    df_f = get_weights(df_fdp)

    oxides = ['SiO2', 'TiO2', 'Al2O3', 'Cr2O3', 'FeO', 'MnO', 'MgO', 'CaO', 'Na2O', 'K2O']

    df_f = _norm_cations_oxygen(df_f, oxides, 5)

    df_f.loc[df_f['NormOxygen_Total'] < 8, 'temp'] = df_f['NormOxygen_Total']
    df_f.loc[2 * (8 - df_f['temp']) < df_f['FE_NormCations'], 'FE3_Atom'] = df_f['FE_NormCations']
    df_f.loc[2 * (8 - df_f['temp']) >= df_f['FE_NormCations'], 'FE3_Atom'] = 8 - df_f['NormOxygen_Total']
    df_f.loc[df_f['NormOxygen_Total'] >= 8, 'FE3_Atom'] = 0
    df_f.pop('temp')
    df_f['FE2_Atom'] = df_f['FE_NormCations'] - df_f['FE3_Atom']

    df_f['anorthite'] = np.round((df_f['CA_Atom'] / (df_f['CA_Atom'] + df_f['K_Atom'] + df_f['NA_Atom'])), 2)
    df_f['albite'] = np.round((df_f['NA_Atom'] / (df_f['CA_Atom'] + df_f['K_Atom'] + df_f['NA_Atom'])), 2)
    df_f['orthoclase'] = np.round((df_f['K_Atom'] / (df_f['CA_Atom'] + df_f['K_Atom'] + df_f['NA_Atom'])), 2)
    df_classification = df_f[['albite', 'anorthite', 'orthoclase']]
    df_f['Classification'] = df_classification.idxmax(axis=1)

    # Formula Feldspar A1T4O8
    # sitio A
    df_f['Formula'] = '('
    df_f.loc[df_f['CA_Atom'] > 0, 'Formula'] += 'Ca' + df_f['CA_Atom'].astype('str')
    df_f.loc[df_f['K_Atom'] > 0, 'Formula'] += 'K' + df_f['K_Atom'].astype('str')
    df_f.loc[df_f['NA_Atom'] > 0, 'Formula'] += 'Na' + df_f['NA_Atom'].astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['K_Atom'] + df_f['NA_Atom'] + df_f['CA_Atom']).round(1).astype('str')

    # sito T - tetraedrico
    df_f['Formula'] += '('
    df_f.loc[df_f['AL_Atom'] > 0, 'Formula'] += 'Al' + df_f['AL_Atom'].astype('str')
    df_f.loc[df_f['SI_Atom'] > 0, 'Formula'] += 'Si' + df_f['SI_Atom'].astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['SI_Atom'] + df_f['AL_Atom']).round(1).astype('str')
    df_f['Formula'] += 'O' + df_f['NormOxygen_Total'].round(1).astype('str')

    # df_f.to_excel('formula_feldspar.xlsx')
    # df_out = df_f[['Formula', 'albite', 'anorthite', 'orthoclase', 'K_Atom', 'NA_Atom',
    #                'CA_Atom', 'SI_Atom', 'AL_Atom']]
    cols = df_f.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    df_out = df_f[cols]
    return df_out


def olivine_formula(df_oli):
    df_f = get_weights(df_oli)
    # Oxides of relevance for Olivines
    oxides = ('SiO2', 'TiO2', 'Al2O3', 'Cr2O3', 'FeO', 'MnO', 'MgO', 'CaO', 'Na2O', 'K2O')

    # df_f = get_weights(df_f)
    df_f = _norm_cations_oxygen(df_f, oxides, 3)
    df_f = _test_Fe(df_f, 8)

    df_f['Forsterite'] = np.round((df_f['MG_Atom'] / (df_f['MN_Atom'] + df_f['FE2_Atom'] + df_f['FE3_Atom']
                                                      + df_f['MG_Atom'] + df_f['CA_Atom'])), 2)

    df_f['Fayalite'] = np.round(((df_f['FE2_Atom'] + df_f['FE3_Atom']) / (df_f['MN_Atom'] + df_f['FE2_Atom'] +
                                                                          df_f['FE3_Atom'] + df_f['MG_Atom'] +
                                                                          df_f['CA_Atom'])), 2)

    df_f['Tephroite'] = np.round((df_f['MN_Atom'] / (df_f['MN_Atom'] + df_f['FE2_Atom'] + df_f['FE3_Atom']
                                                     + df_f['MG_Atom'] + df_f['CA_Atom'])), 2)

    df_f['Ca-Olivine'] = np.round((df_f['CA_Atom'] / (df_f['MN_Atom'] + df_f['FE2_Atom'] + df_f['FE3_Atom']
                                                      + df_f['MG_Atom'] + df_f['CA_Atom'])), 2)

    df_classification = df_f[['Forsterite', 'Fayalite', 'Tephroite', 'Ca-Olivine']]
    df_f['Classification'] = df_classification.idxmax(axis=1)

    # Formula Olivina M2TO4
    # sitio M - octaedrico
    df_f['Formula'] = '('
    df_f.loc[df_f['CA_Atom'] > 0, 'Formula'] += 'Ca' + df_f['CA_Atom'].round(1).astype('str')
    df_f.loc[df_f['MN_Atom'] > 0, 'Formula'] += 'Mn' + df_f['MN_Atom'].round(1).astype('str')
    df_f.loc[df_f['FE2_Atom'] > 0, 'Formula'] += 'Fe' + df_f['FE2_Atom'].round(1).astype('str')
    df_f.loc[df_f['MG_Atom'] > 0, 'Formula'] += 'Mg' + df_f['MG_Atom'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['FE2_Atom'] + df_f['MG_Atom'] +
                                   df_f['CA_Atom'] + df_f['MN_Atom']).round(1).astype('str')

    # sito T - tetraedrico
    df_f.loc[df_f['SI_Atom'] > 0, 'Formula'] += '(Si' + df_f['SI_Atom'].astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['SI_Atom']).round(1).astype('str')

    df_f['Formula'] += 'O' + df_f['NormOxygen_Total'].round(1).astype('str')
    # df_f.to_excel('formula_olivina.xlsx')
    #df_out = df_f[['Formula', 'Forsterite', 'Fayalite', 'Tephroite', 'Ca-Olivine']]
    cols = df_f.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    df_out = df_f[cols]
    return df_out


def pyroxene_formula(df_px):
    df_f = get_weights(df_px)
    oxides = ['SiO2', 'TiO2', 'Al2O3', 'Cr2O3', 'FeO', 'MnO', 'MgO', 'CaO', 'Na2O']
    df_f = _norm_cations_oxygen(df_f, oxides, 4)
    df_f.loc[6 - df_f['NormOxygen_Total'] > 0, 'temp'] = df_f['NormOxygen_Total']
    df_f.loc[2 * (6 - df_f['temp']) < df_f['FE_NormOxygen'], 'FE3_Atom'] = 2 * (6 - df_f['temp'])
    df_f.loc[2 * (6 - df_f['temp']) >= df_f['FE_NormCations'], 'FE3_Atom'] = df_f['FE_NormOxygen']
    df_f.loc[6 - df_f['NormOxygen_Total'] <= 0, 'FE3_Atom'] = 0
    df_f.pop('temp')
    df_f['FE2_Atom'] = df_f['FE_NormCations'] - df_f['FE3_Atom']

    df_f['NMg_FeT'] = (df_f['FE2_Atom'] + df_f['FE3_Atom']) / (df_f['FE2_Atom'] + df_f['FE3_Atom'] + df_f['MG_Atom'])
    df_f['NMg_Fe2'] = (df_f['FE2_Atom'] + df_f['FE3_Atom']) / (df_f['FE2_Atom'] + df_f['FE3_Atom'] + df_f['MG_Atom'])
    df_f['T_Si'] = df_f['SI_Atom']

    df_f.loc[df_f['SI_Atom'] < 2, 'temp'] = df_f['SI_Atom']
    df_f.loc[(2 - df_f['temp']) < df_f['AL_Atom'], 'T_Al'] = 2 - df_f['temp']
    df_f.loc[(2 - df_f['temp']) >= df_f['AL_Atom'], 'T_Al'] = df_f['AL_Atom']
    df_f.loc[df_f['SI_Atom'] >= 2, 'T_Al'] = 0
    df_f.pop('temp')

    df_f['M1_Ti'] = df_f['TI_Atom']

    df_f.loc[df_f['SI_Atom'] + df_f['AL_Atom'] - (df_f['T_Si'] + df_f['T_Al']) == 0,
             'M1_Al'] = 0
    df_f.loc[df_f['SI_Atom'] + df_f['AL_Atom'] - (df_f['T_Si'] + df_f['T_Al']) != 0,
             'M1_Al'] = df_f['SI_Atom'] + df_f['AL_Atom'] - (df_f['T_Si'] + df_f['T_Al'])

    df_f['M1_Cr'] = df_f['CR_Atom']
    df_f['M1_Fe3'] = df_f['FE3_Atom']
    df_f['M2_Ca'] = df_f['CA_Atom']
    df_f['M2_Na'] = df_f['NA_Atom']
    df_f['M2_Mn'] = df_f['MN_Atom']

    df_f.loc[df_f['M2_Ca'] + df_f['M2_Na'] + df_f['M2_Mn'] < 1,
             'M2_Fe2'] = df_f['NMg_Fe2'] * (1 - (df_f['M2_Ca'] + df_f['M2_Na'] + df_f['M2_Mn']))
    df_f.loc[df_f['M2_Ca'] + df_f['M2_Na'] + df_f['M2_Mn'] >= 1, 'M2_Fe2'] = 0

    df_f.loc[df_f['FE2_Atom'] - df_f['M2_Fe2'] > 0, 'M1_Fe2'] = df_f['FE2_Atom'] - df_f['M2_Fe2']
    df_f.loc[df_f['FE2_Atom'] - df_f['M2_Fe2'] <= 0, 'M1_Fe2'] = 0

    df_f.loc[(1 - df_f['NMg_Fe2']) * (1 - (df_f['M2_Ca'] + df_f['M2_Na'] + df_f['M2_Mn'])) > 0,
             'M2_Mg'] = (1 - df_f['NMg_Fe2']) * (1 - (df_f['M2_Ca'] + df_f['M2_Na'] + df_f['M2_Mn']))
    df_f.loc[(1 - df_f['NMg_Fe2']) * (1 - (df_f['M2_Ca'] + df_f['M2_Na'] + df_f['M2_Mn'])) <= 0, 'M2_Mg'] = 0

    df_f.loc[df_f['MG_Atom'] - df_f['M2_Mg'] > 0, 'M1_Mg'] = df_f['MG_Atom'] - df_f['M2_Mg']
    df_f.loc[df_f['MG_Atom'] - df_f['M2_Mg'] <= 0, 'M1_Mg'] = 0

    df_f.loc[df_f['FE2_Atom'] + df_f['CA_Atom'] + df_f['MG_Atom'] > 0,
             'Ortho_sum'] = df_f['FE2_Atom'] + df_f['CA_Atom'] + df_f['MG_Atom']
    df_f['Wollastonite'] = 100 * (df_f['CA_Atom'] / df_f['Ortho_sum'])
    df_f['Wollastonite'] = df_f['Wollastonite'].round(2)

    df_f['Enstatite'] = 100 * (df_f['MG_Atom'] / df_f['Ortho_sum'])
    df_f['Enstatite'] = df_f['Enstatite'].round(2)

    df_f['Ferrosilite'] = 100 * (df_f['FE2_Atom'] / df_f['Ortho_sum'])
    df_f['Ferrosilite'] = df_f['Ferrosilite'].round(2)

    df_f.loc[(df_f['FE3_Atom'] + df_f['NA_Atom'] + df_f['CA_Atom']) > 0,
             'Sodic_sum'] = df_f['FE3_Atom'] + df_f['NA_Atom'] + df_f['CA_Atom']

    df_f['Aegerine'] = np.round(100 * (df_f['FE3_Atom'] / df_f['Sodic_sum']), 2)
    df_f['Jadeite'] = np.round(100 * (df_f['NA_Atom'] / df_f['Sodic_sum']), 2)
    df_f['Diopside'] = np.round(100 * (df_f['CA_Atom'] / df_f['Sodic_sum']), 2)

    df_f['Ortho_Calcic_Total'] = df_f['Wollastonite'] + df_f['Enstatite'] + df_f['Ferrosilite']
    df_f['Sodic_Calcic_Total'] = df_f['Aegerine'] + df_f['Jadeite'] + df_f['Diopside']

    df_classification = df_f[['Wollastonite', 'Enstatite', 'Ferrosilite', 'Aegerine', 'Jadeite', 'Diopside']]
    df_f['Classification'] = df_classification.idxmax(axis=1)
    df_f.loc[df_f['FE3_Atom'] + df_f['NA_Atom'] + df_f['CA_Atom'] == 0, 'temp'] = 0
    # print(df_f['FE3_Atom'] + df_f['NA_Atom'] + df_f['CA_Atom'])
    # print(df_f['Classification'])

    # Formula Piroxenio (M1)(M2)(T)2O6
    # M1 - octaedrico
    # df_f = df_f.round(1)
    df_f['Formula'] = '('
    df_f.loc[df_f['M2_Ca'].round(1) > 0, 'Formula'] += 'Ca' + df_f['M2_Ca'].round(1).astype('str')
    df_f.loc[df_f['M2_Na'].round(1) > 0, 'Formula'] += 'Na' + df_f['M2_Na'].round(1).astype('str')
    df_f.loc[df_f['M2_Mn'].round(1) > 0, 'Formula'] += 'Mn' + df_f['M2_Mn'].round(1).astype('str')
    df_f.loc[df_f['M2_Fe2'].round(1) > 0, 'Formula'] += 'Fe' + df_f['M2_Fe2'].round(1).astype('str')
    df_f.loc[df_f['M2_Mg'].round(1) > 0, 'Formula'] += 'Mg' + df_f['M2_Mg'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['M2_Ca'].round(1) + df_f['M2_Na'].round(1) +
                                   df_f['M2_Mn'].round(1) + df_f['M2_Fe2'].round(1) +
                                   df_f['M2_Mg'].round(1)).round(1).astype('str')
    # M2 octaedrico
    df_f['Formula'] += '('
    df_f.loc[df_f['M1_Ti'].round(1) > 0, 'Formula'] += 'Ti' + df_f['M1_Ti'].round(1).astype('str')
    df_f.loc[df_f['M1_Fe3'].round(1) > 0, 'Formula'] += 'Fe' + df_f['M1_Fe3'].round(1).astype('str')
    df_f.loc[df_f['M1_Al'].round(1) > 0, 'Formula'] += 'Al' + df_f['M1_Al'].round(1).astype('str')
    df_f.loc[df_f['M1_Fe2'].round(1) > 0, 'Formula'] += 'Fe' + df_f['M1_Fe2'].round(1).astype('str')
    df_f.loc[df_f['M1_Mg'].round(1) > 0, 'Formula'] += 'Mg' + df_f['M1_Mg'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['M1_Fe3'].round(1) + df_f['M1_Fe2'].round(1) + df_f['M1_Al'].round(1) +
                                   df_f['M1_Ti'].round(1) + df_f['M1_Mg'].round(1) + df_f['M1_Ti'].round(1)).round(
        1).astype('str')

    # sitio T - tetraedrico
    df_f['Formula'] += '('
    df_f.loc[df_f['T_Al'].round(1) > 0, 'Formula'] += 'Al' + df_f['T_Al'].round(1).astype('str')
    df_f.loc[df_f['T_Si'].round(1) > 0, 'Formula'] += 'Si' + df_f['T_Si'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['T_Si'].round(1) + df_f['T_Al'].round(1)).round(1).astype('str')
    df_f['Formula'] += 'O' + (df_f['NormOxygen_Total']).round(1).astype('str')

    # df_f.to_excel('formula_pyroxene.xlsx')
    df_out = df_f[['Formula', "T_Si", "T_Al", 'M1_Ti', 'M1_Al', 'M1_Cr',
                   'M1_Fe3', 'M2_Ca', 'M2_Na', 'M2_Mn', 'M2_Fe2', 'M1_Fe2',
                   'M2_Mg', 'M1_Mg', 'Wollastonite', 'Enstatite', 'Ferrosilite',
                   'Aegerine', 'Jadeite', 'Diopside']]
    return df_out


def micas_formula(df_mica):
    df_f = get_weights(df_mica)

    df_f.loc[0.287 * df_mica['SIO2'] - 9.552 > 0, 'LI2O'] = 0.287 * df_mica['SIO2'] - 9.552
    df_f.loc[0.287 * df_mica['SIO2'] - 9.552 <= 0, 'LI2O'] = 0
    df_f['LI_mole'] = df_f['LI2O'] / 29.887

    df_f['moles_cations_total'] = df_f.filter(regex='_mole$', axis=1).sum(axis=1)
    df_f['moles_cations_correct'] = df_f['moles_cations_total'] - df_f['CL_mole'] - df_f['F_mole']
    df_f['num_oxygen'] = 22 / df_f['moles_cations_correct']

    # print(df_f)
    oxides = ['SiO2', 'TiO2', 'Al2O3', 'FeO', 'MnO', 'MgO', 'CaO', 'Na2O', 'K2O',
              'F', 'Cl', 'Li', 'Cr']
    for eli in oxides:
        el = eli.split('O')[0]
        if len(el) > 2:
            el = el[:2]
        if el[0] == 'K':
            el = 'K'
        df_f[el.upper() + '_NormCations'] = df_f[el.upper() + '_mole'] * df_f['num_oxygen']

    df_f.loc[4 - (df_f['F_NormCations'] + df_f['CL_NormCations']) > 0,
             'OH'] = 4 - (df_f['F_NormCations'] + df_f['CL_NormCations'])
    df_f.loc[4 - (df_f['F_NormCations'] + df_f['CL_NormCations']) <= 0,
             'OH'] = 0

    df_f['Z_Si'] = df_f['SI_NormCations'] / 2
    df_f.loc[df_f['Z_Si'] > 8, 'Z_Al'] = 0
    df_f.loc[df_f['Z_Si'] <= 8, 'teste'] = df_f['Z_Si']
    df_f.loc[2 * (df_f['AL_NormCations'] / 3) > 8 - df_f['teste'], 'Z_Al'] = 8 - df_f['teste']
    df_f.loc[2 * (df_f['AL_NormCations'] / 3) <= 8 - df_f['teste'], 'Z_Al'] = 2 * (df_f['AL_NormCations'] / 3)
    df_f.pop('teste')

    df_f['Z_sum'] = df_f.filter(regex='^Z_', axis=1).sum(axis=1)

    df_f.loc[df_f['Z_Si'] + 2 * (df_f['AL_NormCations'] / 3) < 8, 'Y_Al'] = 0
    df_f.loc[df_f['Z_Si'] + 2 * (df_f['AL_NormCations'] / 3) >= 8, 'Y_Al'] = 2 * (df_f['AL_NormCations'] / 3) - df_f[
        'Z_Si']
    df_f['Y_Ti'] = df_f['TI_NormCations'] / 2
    df_f['Y_Cr'] = 2 * df_f['CR_NormCations'] / 3
    df_f['Y_Fe'] = df_f['FE_NormCations']
    df_f['Y_Mn'] = df_f['MN_NormCations']
    df_f['Y_Mg'] = df_f['MG_NormCations']
    df_f['Y_Li'] = 2 * df_f['LI_NormCations']
    df_f['Y_sum'] = df_f.filter(regex='^Y_', axis=1).sum(axis=1)

    df_f['X_Ca'] = df_f['CA_NormCations']
    df_f['X_Na'] = 2 * df_f['NA_NormCations']
    df_f['X_K'] = 2 * df_f['K_NormCations']
    df_f['X_sum'] = df_f.filter(regex='^X_', axis=1).sum(axis=1)

    df_f['H_OH'] = df_f['OH']
    df_f['H_F'] = df_f['F_NormCations']
    df_f['H_Cl'] = df_f['CL_NormCations']
    df_f['X_sum'] = df_f.filter(regex='^H_', axis=1).sum(axis=1)

    # Micas - X2Y4-6Z8O2OH4

    # Sitio X
    df_f['Formula'] = '('
    df_f.loc[df_f['X_Ca'].round(1) > 0, 'Formula'] += 'Ca' + df_f['X_Ca'].round(1).astype('str')
    df_f.loc[df_f['X_K'].round(1) > 0, 'Formula'] += 'K' + df_f['X_K'].round(1).astype('str')
    df_f.loc[df_f['X_Na'].round(1) > 0, 'Formula'] += 'Na' + df_f['X_Na'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['X_Ca'].round(1) + df_f['X_Na'].round(1) +
                                   df_f['X_K'].round(1)).round(1).astype('str')
    # sitio Y
    df_f['Formula'] += '('
    df_f.loc[df_f['Y_Ti'].round(1) > 0, 'Formula'] += 'Ti' + df_f['Y_Ti'].round(1).astype('str')
    df_f.loc[df_f['Y_Cr'].round(1) > 0, 'Formula'] += 'Cr' + df_f['Y_Cr'].round(1).astype('str')
    df_f.loc[df_f['Y_Al'].round(1) > 0, 'Formula'] += 'Al' + df_f['Y_Al'].round(1).astype('str')
    # df_f.loc[df_f['Y_Cu'].round(1) > 0, 'Formula'] += 'Cu' + df_f['Y_Cu'].round(1).astype('str')
    # df_f.loc[df_f['Y_Ni'].round(1) > 0, 'Formula'] += 'Ni' + df_f['Y_Ni'].round(1).astype('str')
    df_f.loc[df_f['Y_Mn'].round(1) > 0, 'Formula'] += 'Mn' + df_f['Y_Mn'].round(1).astype('str')
    df_f.loc[df_f['Y_Fe'].round(1) > 0, 'Formula'] += 'Fe' + df_f['Y_Fe'].round(1).astype('str')
    df_f.loc[df_f['Y_Mg'].round(1) > 0, 'Formula'] += 'Mg' + df_f['Y_Mg'].round(1).astype('str')
    df_f.loc[df_f['Y_Li'].round(1) > 0, 'Formula'] += 'Li' + df_f['Y_Li'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['Y_Al'].round(1) + df_f['Y_Ti'].round(1) + df_f['Y_Cr'].round(1) +
                                   df_f['Y_Fe'].round(1) + df_f['Y_Mn'].round(1) + df_f['Y_Mg'].round(1) +
                                   df_f['Y_Li'].round(1)).round(1).astype('str')
    # Sitio Z
    df_f['Formula'] += '('
    df_f.loc[df_f['Z_Al'].round(1) > 0, 'Formula'] += 'Al' + df_f['Z_Al'].round(1).astype('str')
    df_f.loc[df_f['Z_Si'].round(1) > 0, 'Formula'] += 'Si' + df_f['Z_Si'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['Z_Si'].round(1) + df_f['Z_Al'].round(1)).round(1).astype('str')
    # Oxygen
    df_f['Formula'] += 'O' + '20'
    # sitio H
    df_f['Formula'] += '('
    df_f.loc[df_f['H_Cl'].round(1) > 0, 'Formula'] += 'Cl' + df_f['H_Cl'].round(1).astype('str')
    df_f.loc[df_f['H_F'].round(1) > 0, 'Formula'] += 'F' + df_f['H_F'].round(1).astype('str')
    df_f.loc[df_f['H_OH'].round(1) > 0, 'Formula'] += 'OH' + df_f['H_OH'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['H_OH'].round(1) + df_f['H_F'].round(1) +
                                   df_f['H_Cl'].round(1)).round(1).astype('str')

    # df_f.to_excel('formula_mica.xlsx')
    # M2 octaedrico

    # print(df_f)
    cols = df_f.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    df_out = df_f[cols]
    return df_out


def spinel_formula(df_sp):
    oxides = ['SiO2', 'TiO2', 'Al2O3', 'Cr2O3', 'FeOT', 'MnO', 'MgO', 'ZnO']

    df_f = get_weights(df_sp)
    df_f['Total'] = 0
    for ox in oxides:
        ox = ox.upper()
        df_f['Total'] += df_sp[ox]

        el = ox.split('O')[0]
        if len(el) > 2:
            el = el[:2]
        if el[0] == 'K':
            el = 'K'

        df_f[el + '_NormCations'] = 3 * df_f[el + '_mole'] / df_f['moles_cation']
        if el in ['SI', 'TI']:
            df_f[el + '_NormOxygen'] = df_f[el + '_NormCations'] * 2
            df_f[el + '_Atom'] = df_f[el + '_NormCations'].round(1)
        elif el in ['AL', 'CR']:
            df_f[el + '_NormOxygen'] = df_f[el + '_NormCations'] * 3 / 2
            df_f[el + '_Atom'] = df_f[el + '_NormCations'].round(1)
        elif el in ['FE', 'MN', 'MG', 'CA']:
            df_f[el + '_NormOxygen'] = df_f[el + '_NormCations']
            df_f[el + '_Atom'] = df_f[el + '_NormCations'].round(1)
        elif el in ['K', 'NA', 'ZN']:
            df_f[el + '_NormOxygen'] = df_f[el + '_NormCations'] / 2
            df_f[el + '_Atom'] = df_f[el + '_NormCations'].round(1)

    df_f['NormCations_Total'] = df_f.filter(regex='_NormCations$', axis=1).sum(axis=1)
    df_f['NormOxygen_Total'] = df_f.filter(regex='_NormOxygen$', axis=1).sum(axis=1)

    df_f['calculate_charge'] = 4 * (df_f['SI_NormCations'] + df_f['TI_NormCations']) + \
                               3 * (df_f['AL_NormCations'] + df_f['CR_NormCations']) + \
                               2 * (df_f['FE_NormCations'] + df_f['ZN_NormCations']) + \
                               2 * (df_f['MN_NormCations'] + df_f['MG_NormCations'])
    df_f = _test_Fe(df_f, 8)

    df_f['X_Mg'] = df_f['MG_Atom']
    df_f['X_Fe'] = df_f['FE2_Atom'].round(1)
    df_f['X_Zn'] = df_f['ZN_Atom']
    df_f['X_Ti'] = df_f['TI_Atom']
    df_f['X_total'] = df_f.filter(regex='^X_', axis=1).sum(axis=1)

    df_f['Y_Al'] = df_f['AL_Atom']
    df_f['Y_Fe'] = df_f['FE3_Atom'].round(1)
    df_f['Y_Cr'] = df_f['CR_Atom']
    df_f['Y_total'] = df_f.filter(regex='^Y_', axis=1).sum(axis=1)

    # Spinel - XY2O4

    # sitio x
    df_f['Formula'] = '('
    df_f.loc[df_f['X_Ti'] > 0, 'Formula'] += 'Ti' + df_f['X_Ti'].astype('str')
    df_f.loc[df_f['X_Zn'] > 0, 'Formula'] += 'Zn' + df_f['X_Zn'].astype('str')
    df_f.loc[df_f['X_Fe'] > 0, 'Formula'] += 'Fe' + df_f['X_Fe'].astype('str')
    df_f.loc[df_f['X_Mg'] > 0, 'Formula'] += 'Mg' + df_f['X_Mg'].astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['X_total']).round(1).astype('str')
    # sitio Y
    df_f['Formula'] += '('
    df_f.loc[df_f['Y_Cr'] > 0, 'Formula'] += 'Cr' + df_f['Y_Cr'].astype('str')
    df_f.loc[df_f['Y_Fe'] > 0, 'Formula'] += 'Fe' + df_f['Y_Fe'].astype('str')
    df_f.loc[df_f['Y_Al'] > 0, 'Formula'] += 'Al' + df_f['Y_Al'].astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['Y_total']).round(1).astype('str')
    # Oxygen
    df_f['Formula'] += ' O' + df_f['NormOxygen_Total'].round(1).astype('str')

    # df_f.to_excel('formula_spinel.xlsx')
    # print(df_f['Formula'])
    # return df_f[['Formula']]
    cols = df_f.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    df_out = df_f[cols]
    return df_out


def organize_amph(df):

    model = load_regression_model()
    dic = {}
    for i in range(len(df.columns)):
        dic[df.columns[i]] = str(df.columns[i]).strip().upper()
    df = df.rename(columns=dic)

    for i in model['Training_Features']:
        if i not in df.columns:
            df[i] = 0
    return df

def regression_anfibolio(df):
    # Load and run regression Model for a df of Anfibolio Group

    regression_model = load_regression_model()
    keys = regression_model.keys()
    X = df[regression_model['Training_Features']]

    for k in keys:
        if k[-3:] == 'REF':
            df[k] = regression_model[k].predict(X)
    df = df.filter(regex='REF$', axis=1)

    return df


def amphibole_formula(df):

    df_f = regression_anfibolio(organize_amph(df))
   # df_f = pd.concat([df, df_reg], axis=1)

    # Sitio A
    df_f['Formula'] = '('
    df_f.loc[df_f['A.K.REF'].round(1) > 0, 'Formula'] += 'K' + df_f['A.K.REF'].round(1).astype('str')
    df_f.loc[df_f['A.CA.REF'].round(1) > 0, 'Formula'] += 'Ca' + df_f['A.CA.REF'].round(1).astype('str')
    df_f.loc[df_f['A.NA.REF'].round(1) > 0, 'Formula'] += 'Na' + df_f['A.NA.REF'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['A.K.REF'].round(1) + df_f['A.CA.REF'].round(1) +
                                   df_f['A.NA.REF'].round(1)).round(1).astype('str')
    # Sitio B
    df_f['Formula'] += '('
    df_f.loc[df_f['B.CA.REF'].round(1) > 0, 'Formula'] += 'Ca' + df_f['B.CA.REF'].round(1).astype('str')
    df_f.loc[df_f['B.MN2.REF'].round(1) > 0, 'Formula'] += 'Mn' + df_f['B.MN2.REF'].round(1).astype('str')
    df_f.loc[df_f['B.FE2.REF'].round(1) > 0, 'Formula'] += 'Fe' + df_f['B.FE2.REF'].round(1).astype('str')
    df_f.loc[df_f['B.MG.REF'].round(1) > 0, 'Formula'] += 'Mg' + df_f['B.MG.REF'].round(1).astype('str')
    df_f.loc[df_f['B.NA .REF'].round(1) > 0, 'Formula'] += 'Na' + df_f['B.NA .REF'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['B.CA.REF'].round(1) + df_f['B.MN2.REF'].round(1) +
                                   df_f['B.FE2.REF'].round(1) + df_f['B.MG.REF'].round(1) +
                                   df_f['B.NA .REF'].round(1)).round(1).astype('str')

    # Sitio C
    df_f['Formula'] += '('
    df_f.loc[df_f['C.TI.REF'].round(1) > 0, 'Formula'] += 'Ti' + df_f['C.TI.REF'].round(1).astype('str')
    df_f.loc[df_f['C.CR.REF'].round(1) > 0, 'Formula'] += 'Cr' + df_f['C.CR.REF'].round(1).astype('str')
    df_f.loc[df_f['C.FE3.REF'].round(1) > 0, 'Formula'] += 'Fe' + df_f['C.FE3.REF'].round(1).astype('str')
    df_f.loc[df_f['C.AL.REF'].round(1) > 0, 'Formula'] += 'Al' + df_f['C.AL.REF'].round(1).astype('str')
    df_f.loc[df_f['C.NI.REF'].round(1) > 0, 'Formula'] += 'Ni' + df_f['C.NI.REF'].round(1).astype('str')
    df_f.loc[df_f['C.ZN.REF'].round(1) > 0, 'Formula'] += 'Zn' + df_f['C.ZN.REF'].round(1).astype('str')
    df_f.loc[df_f['C.MN2.REF'].round(1) > 0, 'Formula'] += 'Mn' + df_f['C.MN2.REF'].round(1).astype('str')
    df_f.loc[df_f['C.FE2.REF'].round(1) > 0, 'Formula'] += 'Fe' + df_f['C.FE2.REF'].round(1).astype('str')
    df_f.loc[df_f['C.MG.REF'].round(1) > 0, 'Formula'] += 'Mg' + df_f['C.MG.REF'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['C.TI.REF'].round(1) + df_f['C.CR.REF'].round(1) +
                                   df_f['C.FE3.REF'].round(1) + df_f['C.AL.REF'].round(1) +
                                   df_f['C.NI.REF'].round(1) + df_f['C.ZN.REF'].round(1) +
                                   df_f['C.MN2.REF'].round(1) + df_f['C.FE2.REF'].round(1) +
                                   df_f['C.MG.REF'].round(1)).round(1).astype('str')
    # Sitio T
    df_f['Formula'] += '('
    df_f.loc[df_f['T.AL.REF'].round(1) > 0, 'Formula'] += 'Al' + df_f['T.AL.REF'].round(1).astype('str')
    df_f.loc[df_f['T.SI.REF'].round(1) > 0, 'Formula'] += 'Si' + df_f['T.SI.REF'].round(1).astype('str')
    df_f.loc[df_f['T.TI.REF'].round(1) > 0, 'Formula'] += 'Ti' + df_f['T.TI.REF'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['T.AL.REF'].round(1) + df_f['T.SI.REF'].round(1) +
                                   df_f['T.TI.REF'].round(1)).round(1).astype('str')
    # O sum = 22
    df_f['Formula'] += 'O22'

    # Sitio W
    df_f['Formula'] += '('
    df_f['W.O2.REF'] = df_f['W.O.REF'].round(1)
    df_f.loc[1 - df_f['W.O2.REF'] < 0.9, 'Formula'] += 'O' + df_f['W.O.REF'].round(1).astype('str')
    df_f.pop('W.O2.REF')
    df_f.loc[df_f['W.CL.REF'].round(1) > 0, 'Formula'] += 'Cl' + df_f['W.CL.REF'].round(1).astype('str')
    df_f.loc[df_f['W.F.REF'].round(1) > 0, 'Formula'] += 'F' + df_f['W.F.REF'].round(1).astype('str')
    df_f.loc[df_f['W.OH.REF'].round(1) > 0, 'Formula'] += 'OH' + df_f['W.OH.REF'].round(1).astype('str')
    df_f['Formula'] += ')'
    df_f['Formula'] += '\u03A3' + (df_f['W.O.REF'].round(1) + df_f['W.CL.REF'].round(1) +
                                   df_f['W.F.REF'].round(1) + df_f['W.OH.REF'].round(1).round(1)).round(1).astype('str')

    cols = df_f.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    df_out = df_f[cols]

    return df_out


def get_function(group, df_partial):
    #formulas_ready = ['GARNET', 'FELDSPAR', 'OLIVINE', 'PYROXENE', 'MICA', 'SPINEL',
    #                  'AMPHIBOLES']
    formulas_ready = ['GARNET', 'FELDSPAR', 'OLIVINE', 'PYROXENE', 'MICA', 'SPINEL']
    #print('getFUNCTIOn', group)
    if group == formulas_ready[0]:
        df = garnet_formula(df_partial)
    elif group == formulas_ready[1]:
        df = feldspar_formula(df_partial)
    elif group == formulas_ready[2]:
        df = olivine_formula(df_partial)
    elif group == formulas_ready[3]:
        df = pyroxene_formula(df_partial)
    elif group == formulas_ready[4]:
        df = micas_formula(df_partial)
    elif group == formulas_ready[5]:
        df = spinel_formula(df_partial)
    elif group == formulas_ready[6]:
        df = amphibole_formula(df_partial)
    else:
        print(group, 'Not implemented Formula Calulator')
        df = pd.DataFrame(columns=['Formula'])
    return df


def append_df_to_excel(filename, df, sheet_name='Sheet1',
                       startrow=None, truncate_sheet=False,
                       **to_excel_kwargs):
    """
    https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def get_formula(df):
    '''Run formula calculator for groups in list'''
    # import openpyxl
    formulas_ready = ['GARNET', 'FELDSPAR', 'OLIVINE', 'PYROXENE', 'MICA', 'SPINEL',
                      'AMPHIBOLES']
    df = organize(df)

    # groups_df = df['GROUP'].unique()
    groups_df = df['PREDICTED GROUP'].unique()
    print("ALL  GROUPS, ", groups_df)
    dfs = []
    dict = {}
    for group in groups_df:
        print('Calculating Formula for ', group )
        df_partial = df[df['PREDICTED GROUP'] == group]
        # df_partial = df[df['GROUP'] == group]
        df_formula = get_function(group, df_partial)
        if group in formulas_ready:
            dfs.append(pd.concat([df_partial, df_formula['Formula']], axis=1))
            dict[group] = df_formula
        else:
            # dfs.append(df_partial)
            dfs.append(pd.concat([df_partial, df_formula['Formula']], axis=1))
            # append_df_to_excel('formula_calculator_output.xlsx', df_formula, sheet_name=group+'_formula')
            dict[group] = pd.DataFrame()
    df_all = pd.concat(dfs, axis=0)

    return df_all, dict


if __name__ == '__main__':
    df = pd.read_csv('../data_train/Data_imput_SMOTE_Random_Sampler.csv')
    print(df['GROUP'].unique())

    df_grt = df[df['GROUP'] == 'GARNET']
    df_fdp = df[df['GROUP'] == 'FELDSPAR']
    df_oli = df[df['GROUP'] == 'OLIVINE']
    df_px = df[df['GROUP'] == 'PYROXENE']
    df_mc = df[df['GROUP'] == 'MICA']
    df_sp = df[df['GROUP'] == 'SPINEL']
    # garnet_formula(df_grt)
    # olivine_formula(df_oli)
    # feldspar_formula(df_fdp)
    # pyroxene_formula(df_px)
    # micas_formula(df_mc)
    # spinel_formula(df_sp)
    get_formula(df_fdp)
